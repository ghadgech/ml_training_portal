import streamlit as st
import pandas as pd
import random
import os
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openai import OpenAI

st.set_page_config(page_title="ML Training Portal", page_icon="🤖", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
.main-header{background:linear-gradient(135deg,#1e3c72 0%,#2a5298 100%);padding:2rem;border-radius:12px;color:white;text-align:center;margin-bottom:2rem;}
.module-card{background:#f0f4ff;border-left:5px solid #2a5298;padding:1rem 1.5rem;margin:0.4rem 0;border-radius:8px;}
.score-pass{background:linear-gradient(135deg,#11998e 0%,#38ef7d 100%);padding:2rem;border-radius:12px;color:white;text-align:center;}
.score-fail{background:linear-gradient(135deg,#eb3349 0%,#f45c43 100%);padding:2rem;border-radius:12px;color:white;text-align:center;}
</style>
""", unsafe_allow_html=True)

USERS_FILE = "users.csv"
RESULTS_FILE = "results.csv"
RESULTS_FOLDER = "quiz_results"
QUESTIONS_PER_QUIZ = 25
PASS_MARK = 13
os.makedirs(RESULTS_FOLDER, exist_ok=True)

@st.cache_resource
def get_openai_client():
    return OpenAI(api_key=st.secrets["OPENAI_API_KEY"])

# ── PRE-QUIZ BANK (50 questions) ──────────────────────────────────────────────
PRE_QUIZ_QUESTIONS = [
    {"question":"What does ML stand for?","options":["Machine Language","Machine Learning","Model Learning","Multiple Logic"],"answer":"Machine Learning","module":"Basics"},
    {"question":"Machine Learning is a subset of:","options":["Artificial Intelligence","Database Management","Web Development","Networking"],"answer":"Artificial Intelligence","module":"Basics"},
    {"question":"What is supervised learning?","options":["Learning with labeled data","Learning without labels","Learning from rewards","Learning from images only"],"answer":"Learning with labeled data","module":"Basics"},
    {"question":"What is unsupervised learning?","options":["Finding patterns in unlabeled data","Learning with labeled data","Learning from human feedback","Learning from text only"],"answer":"Finding patterns in unlabeled data","module":"Basics"},
    {"question":"Which Python library is mainly used for data manipulation?","options":["Pandas","NumPy","Matplotlib","Seaborn"],"answer":"Pandas","module":"Python"},
    {"question":"Which Python library is used for numerical computation?","options":["NumPy","Pandas","Matplotlib","Plotly"],"answer":"NumPy","module":"Python"},
    {"question":"Which library is used for data visualization?","options":["Matplotlib","NumPy","Pandas","Scikit-learn"],"answer":"Matplotlib","module":"Python"},
    {"question":"What does CSV stand for?","options":["Comma-Separated Values","Computer Science Variables","Column Stored Values","Continuous Sequential Values"],"answer":"Comma-Separated Values","module":"Basics"},
    {"question":"A 'feature' in ML refers to:","options":["An input variable used for prediction","The output prediction","The model itself","A programming function"],"answer":"An input variable used for prediction","module":"Basics"},
    {"question":"The 'target' or 'label' in ML is:","options":["The output we want to predict","An input feature","A data cleaning step","A Python library"],"answer":"The output we want to predict","module":"Basics"},
    {"question":"What is overfitting?","options":["Model performs well on training but poorly on new data","Model performs well on all data","Model that is too simple","Model with no parameters"],"answer":"Model performs well on training but poorly on new data","module":"Basics"},
    {"question":"What is underfitting?","options":["Model is too simple to capture patterns","Model performs very well","Model with too many features","Model trained too long"],"answer":"Model is too simple to capture patterns","module":"Basics"},
    {"question":"Which algorithm predicts continuous numeric values?","options":["Linear Regression","K-Means","Decision Tree Classifier","Naive Bayes"],"answer":"Linear Regression","module":"Algorithms"},
    {"question":"Which algorithm is used for classification tasks?","options":["Logistic Regression","Linear Regression","PCA","K-Means"],"answer":"Logistic Regression","module":"Algorithms"},
    {"question":"What does AI stand for?","options":["Artificial Intelligence","Automatic Input","Algorithmic Interface","Advanced Integration"],"answer":"Artificial Intelligence","module":"Basics"},
    {"question":"What is a DataFrame in Pandas?","options":["A 2D tabular data structure","A 1D array","A Python loop","A visualization tool"],"answer":"A 2D tabular data structure","module":"Python"},
    {"question":"What does pd.read_csv() do?","options":["Reads a CSV file into a DataFrame","Writes data to CSV","Creates a new column","Plots a graph"],"answer":"Reads a CSV file into a DataFrame","module":"Python"},
    {"question":"Which library provides ML algorithms in Python?","options":["Scikit-learn","Seaborn","Plotly","SQLAlchemy"],"answer":"Scikit-learn","module":"Python"},
    {"question":"What does NaN mean in data?","options":["Not a Number (missing value)","Natural Array Node","Negative Applied Node","New Attribute Name"],"answer":"Not a Number (missing value)","module":"Basics"},
    {"question":"What is the purpose of train-test split?","options":["Evaluate model on unseen data","Clean the data","Remove missing values","Add new features"],"answer":"Evaluate model on unseen data","module":"Basics"},
    {"question":"What is data normalization?","options":["Scaling data to a standard range","Removing missing values","Adding new columns","Sorting the data"],"answer":"Scaling data to a standard range","module":"Preprocessing"},
    {"question":"What is the mean of [2, 4, 6, 8, 10]?","options":["6","5","7","8"],"answer":"6","module":"Statistics"},
    {"question":"What is the median of [1, 3, 5, 7, 9]?","options":["5","4","6","7"],"answer":"5","module":"Statistics"},
    {"question":"What is correlation?","options":["Measure of relationship between two variables","A type of regression","A data cleaning method","A Python function"],"answer":"Measure of relationship between two variables","module":"Statistics"},
    {"question":"What is clustering in ML?","options":["Grouping similar data points together","Predicting a target variable","Cleaning missing data","Splitting data into train/test"],"answer":"Grouping similar data points together","module":"Algorithms"},
    {"question":"K-Means is an example of:","options":["Unsupervised learning","Supervised learning","Reinforcement learning","Deep learning"],"answer":"Unsupervised learning","module":"Algorithms"},
    {"question":"What is a decision tree?","options":["A tree-like model of decisions","A type of neural network","A data format","A Python function"],"answer":"A tree-like model of decisions","module":"Algorithms"},
    {"question":"What does accuracy measure in ML?","options":["Percentage of correct predictions","Speed of the model","Size of the dataset","Number of features"],"answer":"Percentage of correct predictions","module":"Evaluation"},
    {"question":"What is one-hot encoding?","options":["Converting categories to binary columns","Normalizing numeric data","Removing outliers","Splitting data"],"answer":"Converting categories to binary columns","module":"Preprocessing"},
    {"question":"What is a confusion matrix used for?","options":["Evaluating classification model performance","Visualizing data distribution","Removing duplicates","Feature selection"],"answer":"Evaluating classification model performance","module":"Evaluation"},
    {"question":"What is deep learning?","options":["ML using multi-layer neural networks","A very thorough data analysis","Learning from textbooks","A supervised algorithm"],"answer":"ML using multi-layer neural networks","module":"Deep Learning"},
    {"question":"What is gradient descent?","options":["Optimization algorithm to minimize loss","A data visualization technique","A type of neural network","A regularization method"],"answer":"Optimization algorithm to minimize loss","module":"Algorithms"},
    {"question":"What is a hyperparameter?","options":["Parameter set before training","A feature of the dataset","An output of the model","A Python variable"],"answer":"Parameter set before training","module":"Evaluation"},
    {"question":"What is cross-validation?","options":["Technique to evaluate model by splitting data multiple ways","A type of deep learning","A data cleaning step","An encoding method"],"answer":"Technique to evaluate model by splitting data multiple ways","module":"Evaluation"},
    {"question":"What is a loss function?","options":["Measures how far predictions are from actual values","A Python error","A feature selection method","A type of layer in neural networks"],"answer":"Measures how far predictions are from actual values","module":"Deep Learning"},
    {"question":"What does EDA stand for?","options":["Exploratory Data Analysis","Enhanced Data Algorithm","External Data Access","Entity Data Archive"],"answer":"Exploratory Data Analysis","module":"Preprocessing"},
    {"question":"Which of the following is NOT a Python data type?","options":["Matrix","List","Dictionary","Tuple"],"answer":"Matrix","module":"Python"},
    {"question":"What does df.shape return in Pandas?","options":["Number of rows and columns","Shape of the first column","Data types of columns","First 5 rows"],"answer":"Number of rows and columns","module":"Python"},
    {"question":"What is the purpose of feature scaling?","options":["Bring features to a similar range","Create new features","Remove features","Visualize features"],"answer":"Bring features to a similar range","module":"Preprocessing"},
    {"question":"Random Forest is an ensemble of:","options":["Decision Trees","Linear Regressions","Neural Networks","K-Means clusters"],"answer":"Decision Trees","module":"Algorithms"},
    {"question":"What is the test size in a typical 80-20 split?","options":["20%","80%","50%","30%"],"answer":"20%","module":"Basics"},
    {"question":"Which pandas function shows the first 5 rows?","options":["df.head()","df.tail()","df.info()","df.describe()"],"answer":"df.head()","module":"Python"},
    {"question":"What is PCA used for?","options":["Dimensionality reduction","Classification","Regression","Data cleaning"],"answer":"Dimensionality reduction","module":"Algorithms"},
    {"question":"What does df.isnull().sum() return?","options":["Count of missing values per column","Sum of all values","Number of columns","Data types"],"answer":"Count of missing values per column","module":"Python"},
    {"question":"Reinforcement learning learns from:","options":["Rewards and penalties","Labeled data","Unlabeled data","Images only"],"answer":"Rewards and penalties","module":"Basics"},
    {"question":"What is the R-squared score used for?","options":["Measuring regression model performance","Measuring classification accuracy","Counting missing values","Feature selection"],"answer":"Measuring regression model performance","module":"Evaluation"},
    {"question":"Standard deviation measures:","options":["Spread of data around the mean","The average value","The maximum value","The most frequent value"],"answer":"Spread of data around the mean","module":"Statistics"},
    {"question":"Which method removes rows with missing values in Pandas?","options":["df.dropna()","df.fillna()","df.remove()","df.delete()"],"answer":"df.dropna()","module":"Python"},
    {"question":"What is the mode in statistics?","options":["Most frequently occurring value","Middle value","Average value","Smallest value"],"answer":"Most frequently occurring value","module":"Statistics"},
    {"question":"Which of these is a classification algorithm?","options":["Naive Bayes","Linear Regression","PCA","K-Means"],"answer":"Naive Bayes","module":"Algorithms"},
]

# ── POST-QUIZ BANK (100 questions, 8 modules) ─────────────────────────────────
POST_QUIZ_QUESTIONS = [
    # Module 1
    {"question":"Which type of ML uses labeled training data?","options":["Supervised Learning","Unsupervised Learning","Reinforcement Learning","Transfer Learning"],"answer":"Supervised Learning","module":"Module 1"},
    {"question":"The bias-variance tradeoff means:","options":["Balancing model complexity to avoid over/underfitting","Choosing between two algorithms","Selecting between features","Optimizing speed vs accuracy"],"answer":"Balancing model complexity to avoid over/underfitting","module":"Module 1"},
    {"question":"What is a training dataset?","options":["Data used to build/fit the model","Data used to test the model","Data used to validate hyperparameters","Raw uncleaned data"],"answer":"Data used to build/fit the model","module":"Module 1"},
    {"question":"Which is an example of a regression problem?","options":["Predicting house prices","Spam vs not-spam email","Identifying cat vs dog","Grouping customers by behavior"],"answer":"Predicting house prices","module":"Module 1"},
    {"question":"Which is an example of a classification problem?","options":["Predicting if a tumor is malignant or benign","Forecasting temperature","Estimating salary from experience","Predicting stock price"],"answer":"Predicting if a tumor is malignant or benign","module":"Module 1"},
    {"question":"Transfer learning involves:","options":["Using knowledge from one task to help with another","Transferring data between databases","Moving models between computers","Copying hyperparameters"],"answer":"Using knowledge from one task to help with another","module":"Module 1"},
    {"question":"The ML pipeline typically includes:","options":["Data → Preprocessing → Model → Evaluation","Model → Data → Output","Training → Testing → Done","Features → Labels only"],"answer":"Data → Preprocessing → Model → Evaluation","module":"Module 1"},
    {"question":"What is a validation set used for?","options":["Tuning hyperparameters during training","Final model evaluation","Data cleaning","Feature engineering"],"answer":"Tuning hyperparameters during training","module":"Module 1"},
    {"question":"Which of these is NOT a type of machine learning?","options":["Compressive Learning","Supervised Learning","Unsupervised Learning","Reinforcement Learning"],"answer":"Compressive Learning","module":"Module 1"},
    {"question":"Regularization in ML is used to:","options":["Prevent overfitting by penalizing complexity","Speed up training","Add more features","Clean data"],"answer":"Prevent overfitting by penalizing complexity","module":"Module 1"},
    {"question":"What is feature engineering?","options":["Creating new informative features from existing data","Removing all features","Selecting the best algorithm","Training the model faster"],"answer":"Creating new informative features from existing data","module":"Module 1"},
    {"question":"Batch learning means:","options":["Training on all available data at once","Training one sample at a time","Training on small random batches","Training then discarding data"],"answer":"Training on all available data at once","module":"Module 1"},
    {"question":"Online learning means:","options":["Learning incrementally from streaming data","Learning from the internet","Training on a cloud server","Training with a web browser"],"answer":"Learning incrementally from streaming data","module":"Module 1"},
    {"question":"Which library provides a unified API for ML in Python?","options":["Scikit-learn","TensorFlow","Keras","PyTorch"],"answer":"Scikit-learn","module":"Module 1"},
    {"question":"What is the curse of dimensionality?","options":["Performance degrades as number of features grows too large","Too much training data causes errors","Model runs out of memory","Too few classes in classification"],"answer":"Performance degrades as number of features grows too large","module":"Module 1"},
    # Module 2
    {"question":"What does numpy.array([1,2,3]).shape return?","options":["(3,)","(1,3)","(3,1)","3"],"answer":"(3,)","module":"Module 2"},
    {"question":"Which pandas method gives statistical summary of a DataFrame?","options":["df.describe()","df.info()","df.head()","df.summary()"],"answer":"df.describe()","module":"Module 2"},
    {"question":"How do you select column 'age' from DataFrame df?","options":["df['age']","df.get('age')","df.age()","df.select('age')"],"answer":"df['age']","module":"Module 2"},
    {"question":"What does df.groupby('city').mean() do?","options":["Groups data by city and calculates mean","Filters rows where city is mean","Sorts by city then mean","Merges city columns"],"answer":"Groups data by city and calculates mean","module":"Module 2"},
    {"question":"Which Matplotlib function creates a scatter plot?","options":["plt.scatter()","plt.plot()","plt.bar()","plt.hist()"],"answer":"plt.scatter()","module":"Module 2"},
    {"question":"What does df.merge() do?","options":["Joins two DataFrames on a common column","Splits a DataFrame","Removes duplicates","Adds new rows"],"answer":"Joins two DataFrames on a common column","module":"Module 2"},
    {"question":"Which NumPy function creates an array of zeros?","options":["np.zeros()","np.empty()","np.blank()","np.null()"],"answer":"np.zeros()","module":"Module 2"},
    {"question":"What does df.pivot_table() create?","options":["A summary table with aggregations","A new DataFrame from CSV","A plot of the data","A copy of the DataFrame"],"answer":"A summary table with aggregations","module":"Module 2"},
    {"question":"Seaborn is built on top of:","options":["Matplotlib","NumPy","Pandas","Plotly"],"answer":"Matplotlib","module":"Module 2"},
    {"question":"What does pd.concat([df1, df2]) do?","options":["Combines two DataFrames vertically","Merges on a key","Joins on index","Subtracts one from the other"],"answer":"Combines two DataFrames vertically","module":"Module 2"},
    {"question":"Which function saves a DataFrame to Excel?","options":["df.to_excel()","df.save_excel()","df.write_excel()","df.export()"],"answer":"df.to_excel()","module":"Module 2"},
    {"question":"What does np.dot(A, B) compute?","options":["Matrix dot product","Element-wise multiplication","Matrix addition","Matrix subtraction"],"answer":"Matrix dot product","module":"Module 2"},
    {"question":"List comprehension [x**2 for x in range(5)] produces:","options":["[0, 1, 4, 9, 16]","[1, 4, 9, 16, 25]","[0, 1, 2, 3, 4]","[0, 2, 4, 6, 8]"],"answer":"[0, 1, 4, 9, 16]","module":"Module 2"},
    {"question":"What does df.fillna(0) do?","options":["Replaces all NaN values with 0","Removes rows with NaN","Counts NaN values","Finds NaN positions"],"answer":"Replaces all NaN values with 0","module":"Module 2"},
    {"question":"Which method sorts a DataFrame by column 'score' descending?","options":["df.sort_values('score', ascending=False)","df.sort('score', reverse=True)","df.order_by('score')","df.arrange('score', desc=True)"],"answer":"df.sort_values('score', ascending=False)","module":"Module 2"},
    # Module 3
    {"question":"Which technique handles missing values by filling with the column mean?","options":["Mean imputation","Mode imputation","Listwise deletion","Hot deck imputation"],"answer":"Mean imputation","module":"Module 3"},
    {"question":"Min-Max scaling transforms data to:","options":["[0, 1] range","[-1, 1] range","Standard normal distribution","Log scale"],"answer":"[0, 1] range","module":"Module 3"},
    {"question":"Z-score standardization transforms data to have:","options":["Mean=0 and Std=1","Mean=1 and Std=0","Values between 0 and 1","Positive values only"],"answer":"Mean=0 and Std=1","module":"Module 3"},
    {"question":"Which scikit-learn class performs Min-Max scaling?","options":["MinMaxScaler","StandardScaler","RobustScaler","Normalizer"],"answer":"MinMaxScaler","module":"Module 3"},
    {"question":"Label encoding converts:","options":["Categories to integer labels","Numbers to categories","Text to one-hot vectors","Categories to floats"],"answer":"Categories to integer labels","module":"Module 3"},
    {"question":"What is an outlier?","options":["A data point far from other observations","A missing value","A duplicate record","A non-numeric column"],"answer":"A data point far from other observations","module":"Module 3"},
    {"question":"The IQR method for outlier detection uses:","options":["Interquartile Range","Mean and Standard Deviation","Min and Max values","Median and Mode"],"answer":"Interquartile Range","module":"Module 3"},
    {"question":"What does a heatmap of the correlation matrix show?","options":["Relationships between all feature pairs","Missing value locations","Distribution of each feature","Time series trends"],"answer":"Relationships between all feature pairs","module":"Module 3"},
    {"question":"What is the purpose of a box plot?","options":["Show data distribution, median, and outliers","Show frequency of values","Show relationship between two variables","Show trend over time"],"answer":"Show data distribution, median, and outliers","module":"Module 3"},
    {"question":"Which encoding is preferred for nominal categories with many levels?","options":["One-Hot Encoding","Label Encoding","Binary Encoding","Target Encoding"],"answer":"One-Hot Encoding","module":"Module 3"},
    {"question":"Feature selection is done to:","options":["Remove irrelevant or redundant features","Add more features","Normalize data","Split the dataset"],"answer":"Remove irrelevant or redundant features","module":"Module 3"},
    {"question":"A histogram shows:","options":["Distribution of a single numeric variable","Relationship between two variables","Category counts","Time series data"],"answer":"Distribution of a single numeric variable","module":"Module 3"},
    {"question":"Data imbalance in classification means:","options":["One class has far more samples than another","Equal number of features and samples","Missing values in target column","Too many features"],"answer":"One class has far more samples than another","module":"Module 3"},
    {"question":"SMOTE is a technique to:","options":["Oversample the minority class synthetically","Remove outliers","Normalize features","Split data"],"answer":"Oversample the minority class synthetically","module":"Module 3"},
    {"question":"What does VIF (Variance Inflation Factor) detect?","options":["Multicollinearity between features","Outliers in data","Missing value patterns","Distribution skewness"],"answer":"Multicollinearity between features","module":"Module 3"},
    # Module 4
    {"question":"In linear regression, the line of best fit minimizes:","options":["Sum of squared residuals","Sum of absolute residuals","Mean of predictions","Maximum error"],"answer":"Sum of squared residuals","module":"Module 4"},
    {"question":"What is RMSE?","options":["Root Mean Squared Error","Relative Mean Standard Error","Random Model Score Evaluation","Regression Model Standard Estimation"],"answer":"Root Mean Squared Error","module":"Module 4"},
    {"question":"An R² value of 0.85 means:","options":["Model explains 85% of variance in target","Model has 85% accuracy","Model error is 15%","Model predicts 85 values correctly"],"answer":"Model explains 85% of variance in target","module":"Module 4"},
    {"question":"Ridge regression adds which penalty to the loss function?","options":["L2 (squared coefficients)","L1 (absolute coefficients)","Elastic Net penalty","No penalty"],"answer":"L2 (squared coefficients)","module":"Module 4"},
    {"question":"Lasso regression adds which penalty?","options":["L1 (absolute value of coefficients)","L2 (squared coefficients)","Elastic net","Dropout penalty"],"answer":"L1 (absolute value of coefficients)","module":"Module 4"},
    {"question":"Polynomial regression is used when:","options":["The relationship between X and Y is non-linear","Data has many outliers","There are too many features","Target variable is categorical"],"answer":"The relationship between X and Y is non-linear","module":"Module 4"},
    {"question":"What does LinearRegression().fit(X_train, y_train) do?","options":["Trains the linear regression model","Predicts values","Evaluates the model","Splits data"],"answer":"Trains the linear regression model","module":"Module 4"},
    {"question":"MAE stands for:","options":["Mean Absolute Error","Model Accuracy Evaluation","Maximum Absolute Error","Mean Adjusted Error"],"answer":"Mean Absolute Error","module":"Module 4"},
    {"question":"Which metric is most sensitive to large errors?","options":["RMSE","MAE","R²","MAPE"],"answer":"RMSE","module":"Module 4"},
    {"question":"Multiple linear regression has:","options":["More than one input feature","More than one target variable","More than one dataset","Multiple model copies"],"answer":"More than one input feature","module":"Module 4"},
    {"question":"The intercept in linear regression (β₀) represents:","options":["Value of Y when all X features are 0","Rate of change of Y with X","Error in the model","Number of features"],"answer":"Value of Y when all X features are 0","module":"Module 4"},
    {"question":"Decision Tree Regressor predicts by:","options":["Averaging target values in leaf nodes","Using a linear equation","Computing probability","Using K nearest neighbors"],"answer":"Averaging target values in leaf nodes","module":"Module 4"},
    # Module 5
    {"question":"In a confusion matrix, True Positives (TP) are:","options":["Correctly predicted positive cases","Incorrectly predicted positive cases","Correctly predicted negative cases","Total positive cases"],"answer":"Correctly predicted positive cases","module":"Module 5"},
    {"question":"Precision is calculated as:","options":["TP / (TP + FP)","TP / (TP + FN)","(TP + TN) / Total","TN / (TN + FP)"],"answer":"TP / (TP + FP)","module":"Module 5"},
    {"question":"Recall (Sensitivity) is calculated as:","options":["TP / (TP + FN)","TP / (TP + FP)","(TP + TN) / Total","TN / (TN + FN)"],"answer":"TP / (TP + FN)","module":"Module 5"},
    {"question":"F1 Score is:","options":["Harmonic mean of Precision and Recall","Average of Precision and Recall","Product of Precision and Recall","Difference of Precision and Recall"],"answer":"Harmonic mean of Precision and Recall","module":"Module 5"},
    {"question":"ROC-AUC measures:","options":["Model's ability to distinguish between classes","Model training speed","Number of features used","Error rate on training data"],"answer":"Model's ability to distinguish between classes","module":"Module 5"},
    {"question":"Support Vector Machine (SVM) works by:","options":["Finding the optimal hyperplane that separates classes","Building multiple decision trees","Computing probabilities for each class","Grouping data into clusters"],"answer":"Finding the optimal hyperplane that separates classes","module":"Module 5"},
    {"question":"K-Nearest Neighbors (KNN) classifies based on:","options":["Majority class among K nearest training points","A learned mathematical equation","Probability distributions","A tree structure"],"answer":"Majority class among K nearest training points","module":"Module 5"},
    {"question":"Naive Bayes is based on:","options":["Bayes' theorem with feature independence assumption","Distance to nearest neighbors","Gradient descent optimization","Tree splitting rules"],"answer":"Bayes' theorem with feature independence assumption","module":"Module 5"},
    {"question":"What does class_weight='balanced' do in scikit-learn?","options":["Adjusts weights to handle class imbalance","Balances number of features","Normalizes input data","Sets equal learning rate"],"answer":"Adjusts weights to handle class imbalance","module":"Module 5"},
    {"question":"Gradient Boosting builds models by:","options":["Sequentially correcting errors of previous models","Training models in parallel","Using random feature subsets only","Averaging multiple independent models"],"answer":"Sequentially correcting errors of previous models","module":"Module 5"},
    {"question":"XGBoost stands for:","options":["Extreme Gradient Boosting","Extended General Boosting","Extra Graph Boost","Exponential Gradient Boost"],"answer":"Extreme Gradient Boosting","module":"Module 5"},
    {"question":"Which metric is best for imbalanced classification problems?","options":["F1 Score or ROC-AUC","Accuracy","RMSE","R²"],"answer":"F1 Score or ROC-AUC","module":"Module 5"},
    # Module 6
    {"question":"K-Means clustering requires you to specify:","options":["Number of clusters (K) in advance","The target variable","Training and test sets","Probability distributions"],"answer":"Number of clusters (K) in advance","module":"Module 6"},
    {"question":"The Elbow Method in K-Means helps determine:","options":["Optimal number of clusters","Optimal learning rate","Best distance metric","Number of features to use"],"answer":"Optimal number of clusters","module":"Module 6"},
    {"question":"PCA reduces dimensions by:","options":["Finding directions of maximum variance","Removing least important features","Clustering features together","Sampling random features"],"answer":"Finding directions of maximum variance","module":"Module 6"},
    {"question":"DBSCAN clustering can identify:","options":["Arbitrarily shaped clusters and noise points","Only spherical clusters","Exactly K clusters","Only 2 clusters"],"answer":"Arbitrarily shaped clusters and noise points","module":"Module 6"},
    {"question":"Hierarchical clustering produces:","options":["A dendrogram showing cluster relationships","Exactly K clusters only","A probability distribution","A regression line"],"answer":"A dendrogram showing cluster relationships","module":"Module 6"},
    {"question":"The Silhouette Score measures:","options":["How well-separated clusters are","Number of clusters","Training loss","Distance to centroid"],"answer":"How well-separated clusters are","module":"Module 6"},
    {"question":"t-SNE is primarily used for:","options":["High-dimensional data visualization","Classification","Regression","Feature engineering"],"answer":"High-dimensional data visualization","module":"Module 6"},
    {"question":"An autoencoder is a neural network for:","options":["Learning compressed representations of data","Classifying images","Predicting sequences","Generating text"],"answer":"Learning compressed representations of data","module":"Module 6"},
    {"question":"In K-Means, the centroid represents:","options":["Mean position of all points in a cluster","The most central data point","The boundary between clusters","The largest value in a cluster"],"answer":"Mean position of all points in a cluster","module":"Module 6"},
    {"question":"Market basket analysis uses which technique?","options":["Association Rule Learning","K-Means Clustering","Linear Regression","Decision Trees"],"answer":"Association Rule Learning","module":"Module 6"},
    # Module 7
    {"question":"K-Fold cross-validation splits data into K equal parts and:","options":["Trains K times, each with a different fold as validation","Trains once on K% of data","Uses K random samples","Tests on K different datasets"],"answer":"Trains K times, each with a different fold as validation","module":"Module 7"},
    {"question":"Grid Search CV exhaustively searches:","options":["All specified hyperparameter combinations","Random hyperparameter combinations","Only 10 combinations","The best learning rate only"],"answer":"All specified hyperparameter combinations","module":"Module 7"},
    {"question":"Random Search CV is preferred over Grid Search when:","options":["Hyperparameter space is large","Dataset is small","Only 2 hyperparameters exist","Speed is not important"],"answer":"Hyperparameter space is large","module":"Module 7"},
    {"question":"L1 regularization (Lasso) can produce:","options":["Sparse models with some coefficients exactly zero","All coefficients shrunk equally","Increased model complexity","Perfect training accuracy"],"answer":"Sparse models with some coefficients exactly zero","module":"Module 7"},
    {"question":"Dropout in neural networks is used to:","options":["Prevent overfitting by randomly disabling neurons","Speed up training","Increase model accuracy on training data","Add more layers"],"answer":"Prevent overfitting by randomly disabling neurons","module":"Module 7"},
    {"question":"Early stopping in model training means:","options":["Stop training when validation loss stops improving","Stop at a fixed number of epochs","Stop when training loss is zero","Stop when accuracy reaches 100%"],"answer":"Stop training when validation loss stops improving","module":"Module 7"},
    {"question":"Stratified K-Fold ensures:","options":["Each fold has same class distribution as full dataset","Folds are equal in size","Folds are randomly ordered","Training set is always largest"],"answer":"Each fold has same class distribution as full dataset","module":"Module 7"},
    {"question":"Bayesian Optimization for hyperparameter tuning:","options":["Intelligently searches using results of previous trials","Random search with uniform distribution","Grid search with Bayes theorem","Exhaustive search method"],"answer":"Intelligently searches using results of previous trials","module":"Module 7"},
    {"question":"Model ensembling combines multiple models to:","options":["Improve prediction accuracy","Reduce training time","Simplify the model","Remove outliers"],"answer":"Improve prediction accuracy","module":"Module 7"},
    {"question":"The learning curve shows:","options":["Training and validation performance vs training size","Gradient descent progress","Feature importance","Hyperparameter sensitivity"],"answer":"Training and validation performance vs training size","module":"Module 7"},
    {"question":"Feature importance in Random Forest measures:","options":["How much each feature reduces impurity","Correlation of features","Number of times feature is selected","Average feature value"],"answer":"How much each feature reduces impurity","module":"Module 7"},
    # Module 8
    {"question":"A neural network layer with no activation function is equivalent to:","options":["A linear transformation","A sigmoid function","A convolution operation","A pooling layer"],"answer":"A linear transformation","module":"Module 8"},
    {"question":"ReLU activation function outputs:","options":["max(0, x)","1/(1+e^-x)","tanh(x)","(e^x - e^-x)/(e^x + e^-x)"],"answer":"max(0, x)","module":"Module 8"},
    {"question":"Backpropagation in neural networks computes:","options":["Gradients of loss with respect to weights","Forward pass predictions","Activation function values","Batch normalization"],"answer":"Gradients of loss with respect to weights","module":"Module 8"},
    {"question":"CNN (Convolutional Neural Networks) are primarily used for:","options":["Image recognition and computer vision","Time series prediction","Text generation","Tabular data classification"],"answer":"Image recognition and computer vision","module":"Module 8"},
    {"question":"RNN (Recurrent Neural Networks) are designed for:","options":["Sequential data like text and time series","Image data","Tabular data","Unsupervised learning only"],"answer":"Sequential data like text and time series","module":"Module 8"},
    {"question":"LSTM solves the problem of:","options":["Vanishing gradient in standard RNNs","Slow training in CNNs","Overfitting in decision trees","Class imbalance"],"answer":"Vanishing gradient in standard RNNs","module":"Module 8"},
    {"question":"The softmax function is used in the output layer for:","options":["Multi-class classification","Binary classification only","Regression problems","Clustering tasks"],"answer":"Multi-class classification","module":"Module 8"},
    {"question":"Batch normalization in deep learning helps to:","options":["Stabilize and accelerate training","Add more layers","Remove outliers","Select features"],"answer":"Stabilize and accelerate training","module":"Module 8"},
    {"question":"The Adam optimizer combines:","options":["Momentum and RMSProp adaptive learning rates","SGD and Grid Search","L1 and L2 regularization","Dropout and Batch Normalization"],"answer":"Momentum and RMSProp adaptive learning rates","module":"Module 8"},
    {"question":"Transfer learning in deep learning typically involves:","options":["Using pre-trained models and fine-tuning on new data","Training from scratch always","Using only the last layer of a model","Copying weights without any modification"],"answer":"Using pre-trained models and fine-tuning on new data","module":"Module 8"},
]

# ── COURSE CONTENT ────────────────────────────────────────────────────────────
COURSE_MODULES = {
    "Module 1: Introduction to Machine Learning": {
        "icon": "🤖",
        "overview": "Understand the fundamentals of Machine Learning, its types, and real-world applications.",
        "topics": [
            "What is Machine Learning? Differences from traditional programming",
            "Types of ML: Supervised, Unsupervised, Reinforcement Learning",
            "The ML Pipeline: Data → Preprocessing → Model → Evaluation → Deployment",
            "Overfitting vs Underfitting — Bias-Variance Tradeoff",
            "Batch learning vs Online learning",
            "Key terminology: Features, Labels, Training/Test/Validation Sets",
            "Real-world ML applications: Healthcare, Finance, Retail, Manufacturing",
        ],
        "tools": ["Python", "Scikit-learn", "Jupyter Notebook"],
        "outcomes": ["Define ML and its types", "Describe the ML workflow", "Identify appropriate ML approach for a problem"],
    },
    "Module 2: Python for Data Science": {
        "icon": "🐍",
        "overview": "Master Python libraries essential for data manipulation, analysis, and visualization.",
        "topics": [
            "Python basics: Lists, Dicts, Functions, Classes",
            "NumPy: Arrays, matrix operations, vectorized computations",
            "Pandas: DataFrames, Series, reading/writing data",
            "Data manipulation: Filtering, groupby, merge, pivot tables",
            "Matplotlib: Line plots, bar charts, histograms, scatter plots",
            "Seaborn: Statistical visualizations, heatmaps, pair plots",
            "Plotly: Interactive visualizations",
        ],
        "tools": ["Python 3.x", "NumPy", "Pandas", "Matplotlib", "Seaborn"],
        "outcomes": ["Write Python code for data manipulation", "Create informative visualizations", "Perform data aggregation and joins"],
    },
    "Module 3: Data Preprocessing & EDA": {
        "icon": "🔧",
        "overview": "Learn to clean, transform, and explore data to prepare it for machine learning.",
        "topics": [
            "Exploratory Data Analysis (EDA) — understand data structure and distribution",
            "Handling missing values: Imputation (mean, median, mode), deletion",
            "Detecting and treating outliers: IQR method, Z-score",
            "Feature scaling: Min-Max normalization, Z-score standardization",
            "Encoding categorical variables: Label Encoding, One-Hot Encoding",
            "Feature selection: Correlation analysis, VIF, Chi-square test",
            "Handling class imbalance: Oversampling, Undersampling, SMOTE",
        ],
        "tools": ["Pandas", "Scikit-learn", "Seaborn", "Matplotlib"],
        "outcomes": ["Perform EDA on real datasets", "Handle missing values and outliers", "Apply encoding and scaling techniques"],
    },
    "Module 4: Supervised Learning — Regression": {
        "icon": "📈",
        "overview": "Predict continuous numeric values using regression algorithms.",
        "topics": [
            "Simple and Multiple Linear Regression",
            "Polynomial Regression for non-linear relationships",
            "Ridge Regression (L2 regularization)",
            "Lasso Regression (L1 regularization) — built-in feature selection",
            "Elastic Net Regression",
            "Decision Tree Regressor and Random Forest Regressor",
            "Evaluation metrics: MAE, MSE, RMSE, R², Adjusted R²",
            "Interpreting model coefficients and feature importance",
        ],
        "tools": ["Scikit-learn", "Pandas", "NumPy", "Matplotlib"],
        "outcomes": ["Build and evaluate regression models", "Choose appropriate regression technique", "Interpret regression metrics"],
    },
    "Module 5: Supervised Learning — Classification": {
        "icon": "🏷️",
        "overview": "Classify data into categories using powerful classification algorithms.",
        "topics": [
            "Logistic Regression for binary and multi-class classification",
            "Decision Trees: Splitting criteria (Gini, Entropy)",
            "Random Forest: Ensemble of decision trees",
            "Support Vector Machines (SVM): Linear and kernel methods",
            "K-Nearest Neighbors (KNN)",
            "Naive Bayes: Gaussian, Multinomial",
            "Gradient Boosting: XGBoost, LightGBM",
            "Evaluation: Accuracy, Precision, Recall, F1, ROC-AUC, Confusion Matrix",
        ],
        "tools": ["Scikit-learn", "XGBoost", "Matplotlib"],
        "outcomes": ["Implement classification algorithms", "Evaluate classifiers using multiple metrics", "Handle imbalanced datasets"],
    },
    "Module 6: Unsupervised Learning": {
        "icon": "🔍",
        "overview": "Discover hidden patterns and structures in unlabeled data.",
        "topics": [
            "K-Means Clustering: Algorithm, Elbow Method, choosing K",
            "Hierarchical Clustering: Agglomerative, Dendrogram",
            "DBSCAN: Density-based clustering, handling noise",
            "Principal Component Analysis (PCA): Dimensionality reduction",
            "t-SNE: High-dimensional data visualization",
            "Association Rule Learning: Apriori, FP-Growth, Market Basket Analysis",
            "Autoencoders for representation learning",
            "Silhouette Score for cluster quality evaluation",
        ],
        "tools": ["Scikit-learn", "Matplotlib", "Seaborn"],
        "outcomes": ["Apply clustering algorithms", "Perform dimensionality reduction", "Interpret cluster quality metrics"],
    },
    "Module 7: Model Evaluation & Hyperparameter Tuning": {
        "icon": "⚙️",
        "overview": "Optimize model performance through rigorous evaluation and hyperparameter tuning.",
        "topics": [
            "Cross-validation: K-Fold, Stratified K-Fold, Leave-One-Out",
            "Grid Search CV: Exhaustive hyperparameter search",
            "Random Search CV: Efficient random hyperparameter search",
            "Bayesian Optimization for smart hyperparameter search",
            "Regularization: L1, L2, Elastic Net, Dropout",
            "Learning curves: Diagnosing bias and variance",
            "Feature importance and model interpretability (SHAP)",
            "Model ensembling: Bagging, Boosting, Stacking",
        ],
        "tools": ["Scikit-learn", "Optuna", "SHAP", "Matplotlib"],
        "outcomes": ["Apply cross-validation correctly", "Tune hyperparameters systematically", "Interpret and improve model performance"],
    },
    "Module 8: Deep Learning & Neural Networks": {
        "icon": "🧠",
        "overview": "Build powerful deep learning models for complex tasks using neural networks.",
        "topics": [
            "Neural Network fundamentals: Neurons, layers, weights, biases",
            "Activation functions: ReLU, Sigmoid, Softmax, Tanh",
            "Backpropagation and gradient descent",
            "CNNs (Convolutional Neural Networks) for image processing",
            "RNNs, LSTMs for sequential data and NLP",
            "Batch Normalization, Dropout for regularization",
            "Optimizers: SGD, Adam, RMSProp",
            "Transfer Learning: Using pre-trained models (VGG, ResNet, BERT)",
            "Introduction to TensorFlow/Keras and PyTorch",
        ],
        "tools": ["TensorFlow", "Keras", "PyTorch", "GPU Computing"],
        "outcomes": ["Build and train neural networks", "Apply CNNs and RNNs to real problems", "Use transfer learning effectively"],
    },
}

# ── HELPERS ───────────────────────────────────────────────────────────────────
def login_user(username, password):
    try:
        users = pd.read_csv(USERS_FILE)
        match = users[(users["username"] == username) & (users["password"] == password)]
        return match.iloc[0].to_dict() if not match.empty else None
    except:
        return None

def is_instructor(username):
    return str(username).lower() in ["instructor", "admin", "teacher"]

def get_score_comment(score, total):
    pct = (score / total) * 100
    if pct >= 92:
        return "Outstanding", "Excellent! You have demonstrated mastery of Machine Learning concepts."
    elif pct >= 80:
        return "Very Good", "Great understanding of ML concepts. Keep up the excellent work!"
    elif pct >= 68:
        return "Good", "Solid foundation. Review a few topics to strengthen your knowledge further."
    elif pct >= 52:
        return "Pass", "You have passed. Continue practicing to deepen your ML understanding."
    else:
        return "Needs Improvement", "Please review the course materials carefully and retake the quiz."

def get_excel_path(username):
    return os.path.join(RESULTS_FOLDER, f"{username}_quiz_results.xlsx")

def save_to_excel(username, name, batch, quiz_type, score, total, questions, answers):
    filepath = get_excel_path(username)
    pct = round((score / total) * 100, 1)
    grade, comment = get_score_comment(score, total)
    passed = "Pass" if score >= PASS_MARK else "Fail"
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Summary sheet
    if "Summary" not in wb.sheetnames:
        ws = wb.create_sheet("Summary", 0)
        hdrs = ["Attempt", "Date", "Name", "Batch", "Quiz Type", "Score", "Total", "Percentage", "Result", "Grade", "Comments"]
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", fgColor="1e3c72")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        col_widths = [8, 20, 20, 12, 12, 8, 8, 12, 10, 15, 55]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        ws = wb["Summary"]

    attempt_num = ws.max_row  # row 1 = header
    row_data = [attempt_num, now, name, batch, quiz_type, score, total, f"{pct}%", passed, grade, comment]
    ws.append(row_data)
    fill = "d4edda" if passed == "Pass" else "f8d7da"
    for c in range(1, 12):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Detail sheet
    sname = f"Att{attempt_num}_{quiz_type[:3]}"[:31]
    wd = wb.create_sheet(sname)
    wd.merge_cells("A1:F1")
    wd["A1"] = f"Attempt #{attempt_num} | {quiz_type} | {name} ({batch}) | {now}"
    wd["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    wd["A1"].fill = PatternFill("solid", fgColor="1e3c72")
    wd["A1"].alignment = Alignment(horizontal="center")
    wd.merge_cells("A2:F2")
    wd["A2"] = f"Score: {score}/{total} ({pct}%) | {passed} | {grade}: {comment}"
    wd["A2"].font = Font(bold=True, size=11)
    wd["A2"].fill = PatternFill("solid", fgColor="d4edda" if passed == "Pass" else "f8d7da")
    wd["A2"].alignment = Alignment(horizontal="center")

    for c, h in enumerate(["#", "Module", "Question", "Your Answer", "Correct Answer", "Result"], 1):
        cell = wd.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2a5298")
        cell.alignment = Alignment(horizontal="center")
    for w, col in zip([5, 12, 50, 35, 35, 12], range(1, 7)):
        wd.column_dimensions[get_column_letter(col)].width = w

    for i, q in enumerate(questions):
        ua = answers.get(i, "Not Answered")
        correct = q["answer"]
        ok = ua == correct
        wd.append([i+1, q.get("module",""), q["question"], ua, correct, "✓ Correct" if ok else "✗ Wrong"])
        r = wd.max_row
        fc = "d4edda" if ok else "f8d7da"
        for c in range(1, 7):
            cell = wd.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor=fc)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        wd.row_dimensions[r].height = 38

    wb.save(filepath)
    return filepath

def generate_certificate(name, batch, quiz_type, score, total, date_str):
    pct = round((score / total) * 100, 1)
    grade, comment = get_score_comment(score, total)
    passed = score >= PASS_MARK
    score_color = "#11998e" if passed else "#eb3349"
    result_label = "PASS" if passed else "FAIL"

    fig, ax = plt.subplots(figsize=(14, 10))
    ax.set_xlim(0, 14); ax.set_ylim(0, 10); ax.axis("off")
    fig.patch.set_facecolor("#f8f9ff"); ax.set_facecolor("#f8f9ff")

    # Borders
    ax.add_patch(patches.FancyBboxPatch((0.15,0.15),13.7,9.7, boxstyle="round,pad=0.1",
        linewidth=4, edgecolor="#1e3c72", facecolor="white", zorder=0))
    ax.add_patch(patches.FancyBboxPatch((0.35,0.35),13.3,9.3, boxstyle="round,pad=0.05",
        linewidth=2, edgecolor="#c9a227", facecolor="none", zorder=1))

    # Header
    ax.add_patch(patches.Rectangle((0.15,8.2),13.7,1.65, facecolor="#1e3c72", zorder=2))
    ax.text(7,9.3,"ML TRAINING PORTAL", ha="center", va="center", fontsize=22,
            fontweight="bold", color="white", zorder=3)
    ax.text(7,8.7,"Certificate of Completion", ha="center", va="center",
            fontsize=16, color="#c9a227", style="italic", zorder=3)

    # Body
    ax.text(7,7.6,"This is to certify that", ha="center", va="center", fontsize=13, color="#555555", zorder=3)
    ax.text(7,6.9, name, ha="center", va="center", fontsize=28, fontweight="bold", color="#1e3c72", zorder=3)
    ax.plot([3.2,10.8],[6.57,6.57], color="#c9a227", linewidth=2, zorder=3)
    ax.text(7,6.1,f"from {batch} has successfully completed", ha="center", va="center", fontsize=13, color="#555555", zorder=3)
    ax.text(7,5.45,f"Machine Learning {quiz_type}", ha="center", va="center",
            fontsize=17, fontweight="bold", color="#2a5298", zorder=3)

    # Score
    ax.add_patch(patches.FancyBboxPatch((4.3,4.15),5.4,0.95, boxstyle="round,pad=0.1",
        facecolor=score_color, alpha=0.15, zorder=2))
    ax.text(7,4.63,f"Score: {score} / {total}   ({pct}%)", ha="center", va="center",
            fontsize=15, fontweight="bold", color=score_color, zorder=3)

    # Grade
    ax.add_patch(patches.FancyBboxPatch((5.3,3.25),3.4,0.72, boxstyle="round,pad=0.1",
        facecolor=score_color, zorder=2))
    ax.text(7,3.61,f"Grade: {grade}", ha="center", va="center",
            fontsize=13, fontweight="bold", color="white", zorder=3)

    ax.text(7,2.78,f'"{comment}"', ha="center", va="center",
            fontsize=11, color="#555555", style="italic", zorder=3)

    # Footer
    ax.plot([0.5,13.5],[1.85,1.85], color="#1e3c72", linewidth=1, alpha=0.3, zorder=3)
    for x, label, val in [(2.5,"Date of Completion", date_str),(7,"Program","Machine Learning Training"),(11.5,"Result", result_label)]:
        ax.text(x,1.45, label, ha="center", va="center", fontsize=9, color="#888888", zorder=3)
        ax.text(x,1.1, val, ha="center", va="center", fontsize=12 if x!=11.5 else 16,
                fontweight="bold", color="#333333" if x!=11.5 else score_color, zorder=3)
        ax.plot([x-1.5, x+1.5],[0.88,0.88], color="#333333", linewidth=1, zorder=3)

    # Corner circles
    for cx, cy in [(0.65,9.55),(13.35,9.55),(0.65,0.45),(13.35,0.45)]:
        for r, a in [(0.28,0.3),(0.2,0.5),(0.12,0.8)]:
            ax.add_patch(plt.Circle((cx,cy), r, color="#c9a227", fill=False, alpha=a, zorder=3))

    buf = BytesIO()
    plt.tight_layout(pad=0)
    plt.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="#f8f9ff")
    plt.close()
    buf.seek(0)
    return buf.read()

# ── SESSION STATE INIT ────────────────────────────────────────────────────────
for k, v in {"logged_in":False,"username":"","name":"","batch":"",
             "quiz_active":False,"quiz_type":"","quiz_questions":[],
             "q_idx":0,"user_answers":{},"quiz_submitted":False,"quiz_score":0}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── QUIZ HELPERS ──────────────────────────────────────────────────────────────
def start_quiz(quiz_type):
    bank = PRE_QUIZ_QUESTIONS if quiz_type == "Pre Quiz" else POST_QUIZ_QUESTIONS
    st.session_state.quiz_active = True
    st.session_state.quiz_type = quiz_type
    st.session_state.quiz_questions = random.sample(bank, min(QUESTIONS_PER_QUIZ, len(bank)))
    st.session_state.q_idx = 0
    st.session_state.user_answers = {}
    st.session_state.quiz_submitted = False
    st.session_state.quiz_score = 0

def submit_quiz():
    qs = st.session_state.quiz_questions
    ans = st.session_state.user_answers
    score = sum(1 for i, q in enumerate(qs) if ans.get(i) == q["answer"])
    st.session_state.quiz_score = score
    st.session_state.quiz_submitted = True
    save_to_excel(st.session_state.username, st.session_state.name, st.session_state.batch,
                  st.session_state.quiz_type, score, len(qs), qs, ans)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    pct = round((score / len(qs)) * 100, 1)
    row = pd.DataFrame({"username":[st.session_state.username],"name":[st.session_state.name],
                        "batch":[st.session_state.batch],"quiz_type":[st.session_state.quiz_type],
                        "score":[score],"total":[len(qs)],"percentage":[pct],
                        "result":["Pass" if score >= PASS_MARK else "Fail"],"date":[now]})
    try:
        old = pd.read_csv(RESULTS_FILE)
        pd.concat([old, row], ignore_index=True).to_csv(RESULTS_FILE, index=False)
    except:
        row.to_csv(RESULTS_FILE, index=False)

# ── PAGE: LOGIN ───────────────────────────────────────────────────────────────
def show_login():
    st.markdown('<div class="main-header"><h1>🤖 ML Training Portal</h1><p>Machine Learning Certification Program</p></div>', unsafe_allow_html=True)
    _, col, _ = st.columns([1, 1.2, 1])
    with col:
        st.markdown("### Login to Your Portal")
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        if st.button("🔐 Login", use_container_width=True, type="primary"):
            user = login_user(username, password)
            if user:
                for k in ["logged_in","username","name","batch"]:
                    st.session_state[k] = [True, user["username"], user["name"], user["batch"]][["logged_in","username","name","batch"].index(k)]
                st.session_state.logged_in = True
                st.session_state.username = user["username"]
                st.session_state.name = user["name"]
                st.session_state.batch = user["batch"]
                st.rerun()
            else:
                st.error("Invalid username or password.")

# ── PAGE: HOME ────────────────────────────────────────────────────────────────
def show_home():
    st.markdown(f'<div class="main-header"><h1>🤖 ML Training Portal</h1><p>Welcome, <b>{st.session_state.name}</b> | {st.session_state.batch}</p></div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    try:
        wb = openpyxl.load_workbook(get_excel_path(st.session_state.username))
        attempts = wb["Summary"].max_row - 1
    except:
        attempts = 0
    c1.metric("Total Modules", "8")
    c2.metric("Questions / Quiz", "25 random")
    c3.metric("Pass Mark", f"{PASS_MARK}/25 (52%)")
    c4.metric("Your Attempts", attempts)
    st.markdown("---")
    st.markdown("### 🗂️ Your 8-Module Curriculum")
    cols = st.columns(4)
    for i, (name, data) in enumerate(COURSE_MODULES.items()):
        with cols[i % 4]:
            st.markdown(f'<div class="module-card"><b>{data["icon"]} {name.split(":")[0]}</b><br><small>{name.split(":",1)[1].strip()}</small></div>', unsafe_allow_html=True)
    st.markdown("---")
    a, b = st.columns(2)
    a.info("**📝 Pre Quiz** — 25 random baseline questions. Take before training starts.")
    b.success("**✅ Post Quiz** — 25 random advanced questions covering all 8 modules.")

# ── PAGE: COURSE CONTENT ──────────────────────────────────────────────────────
def show_course_content():
    st.markdown("## 📚 Course Content — 8 Modules")
    for mod_name, data in COURSE_MODULES.items():
        with st.expander(f"{data['icon']} {mod_name}", expanded=False):
            st.markdown(f"**Overview:** {data['overview']}")
            st.markdown("**Topics Covered:**")
            for t in data["topics"]:
                st.markdown(f"  - {t}")
            col1, col2 = st.columns(2)
            with col1:
                st.markdown(f"**Tools:** {' • '.join(data['tools'])}")
            with col2:
                st.markdown("**Learning Outcomes:**")
                for o in data["outcomes"]:
                    st.markdown(f"  ✓ {o}")

# ── PAGE: QUIZ ────────────────────────────────────────────────────────────────
def show_quiz_page(quiz_type):
    # Start screen
    if not st.session_state.quiz_active or st.session_state.quiz_type != quiz_type:
        st.markdown(f"## {'📝' if quiz_type=='Pre Quiz' else '✅'} {quiz_type}")
        desc = "Assesses your **baseline ML knowledge** before training." if quiz_type == "Pre Quiz" else "Tests your mastery across **all 8 training modules**."
        (st.info if quiz_type == "Pre Quiz" else st.success)(desc)
        _, mid, _ = st.columns([1, 1, 1])
        with mid:
            st.markdown(f"""
| Detail | Value |
|--------|-------|
| Questions | 25 (randomly selected) |
| Options per question | 4 MCQ |
| Pass Mark | {PASS_MARK}/25 (52%) |
| Time Limit | None |
| Attempts | Unlimited |
""")
            if st.button(f"▶️ Start {quiz_type}", use_container_width=True, type="primary"):
                start_quiz(quiz_type)
                st.rerun()
        return

    if st.session_state.quiz_submitted:
        show_results()
        return

    # Active quiz
    qs = st.session_state.quiz_questions
    idx = st.session_state.q_idx
    total = len(qs)
    q = qs[idx]

    st.markdown(f"## {'📝' if quiz_type=='Pre Quiz' else '✅'} {quiz_type}")
    st.progress((idx + 1) / total)
    answered = len(st.session_state.user_answers)
    st.caption(f"Question {idx+1} of {total}  •  Answered: {answered}/{total}  •  {total-answered} remaining")

    st.markdown(f"---\n### Q{idx+1}. {q['question']}")
    st.caption(f"Topic: {q.get('module','')}")

    opts = q["options"]
    prev = st.session_state.user_answers.get(idx, None)
    def_idx = opts.index(prev) if prev in opts else 0
    chosen = st.radio("Select your answer:", opts, index=def_idx, key=f"r_{idx}_{quiz_type}")
    st.session_state.user_answers[idx] = chosen

    st.markdown("---")
    c1, c2, _, c4, c5 = st.columns([1, 1, 2, 1.2, 1.2])
    with c1:
        if idx > 0 and st.button("◀ Prev", use_container_width=True):
            st.session_state.q_idx -= 1; st.rerun()
    with c2:
        if idx < total - 1 and st.button("Next ▶", use_container_width=True, type="primary"):
            st.session_state.q_idx += 1; st.rerun()
    with c4:
        jump = st.selectbox("Jump to Q:", range(1, total+1), index=idx, label_visibility="collapsed")
        if jump - 1 != idx:
            st.session_state.q_idx = jump - 1; st.rerun()
    with c5:
        unanswered = total - len(st.session_state.user_answers)
        if unanswered > 0:
            st.caption(f"⚠️ {unanswered} unanswered")
        if st.button("✅ Submit Quiz", use_container_width=True, type="primary"):
            submit_quiz(); st.rerun()

    with st.expander("📋 Question Navigator", expanded=False):
        cols5 = st.columns(5)
        for i in range(total):
            with cols5[i % 5]:
                icon = "🟢" if i in st.session_state.user_answers and i != idx else ("🔵" if i == idx else "⚪")
                if st.button(f"{icon} Q{i+1}", key=f"nav_{i}_{quiz_type}"):
                    st.session_state.q_idx = i; st.rerun()

# ── PAGE: RESULTS ─────────────────────────────────────────────────────────────
def show_results():
    qs = st.session_state.quiz_questions
    ans = st.session_state.user_answers
    score = st.session_state.quiz_score
    total = len(qs)
    pct = round((score / total) * 100, 1)
    passed = score >= PASS_MARK
    grade, comment = get_score_comment(score, total)
    quiz_type = st.session_state.quiz_type
    date_str = datetime.now().strftime("%B %d, %Y")

    card = "score-pass" if passed else "score-fail"
    st.markdown(f"""<div class="{card}">
        <h2>{'PASS ✓' if passed else 'FAIL ✗'}</h2>
        <h1 style="font-size:3rem">{score} / {total}</h1>
        <h3>{pct}% — {grade}</h3>
        <p><i>{comment}</i></p>
    </div>""", unsafe_allow_html=True)
    st.markdown("")

    c1, c2, c3 = st.columns(3)
    with c1:
        cert = generate_certificate(st.session_state.name, st.session_state.batch, quiz_type, score, total, date_str)
        st.download_button("🏆 Download Certificate (PNG)", cert,
            f"Certificate_{st.session_state.username}_{quiz_type.replace(' ','_')}.png", "image/png", use_container_width=True)
    with c2:
        excel_path = get_excel_path(st.session_state.username)
        if os.path.exists(excel_path):
            with open(excel_path, "rb") as f:
                st.download_button("📊 Download Results (Excel)", f.read(),
                    f"{st.session_state.username}_quiz_results.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with c3:
        if st.button("🔄 Start New Quiz", use_container_width=True, type="primary"):
            st.session_state.quiz_active = False
            st.session_state.quiz_submitted = False
            st.rerun()

    st.markdown("---\n### 📋 Detailed Question Review")
    correct_ct = sum(1 for i, q in enumerate(qs) if ans.get(i) == q["answer"])
    st.markdown(f"**{correct_ct} correct  •  {total - correct_ct} wrong  •  {total} total**")

    for i, q in enumerate(qs):
        ua = ans.get(i, "Not Answered")
        ok = ua == q["answer"]
        color = "#11998e" if ok else "#eb3349"
        bg = "#f0fff4" if ok else "#fff5f5"
        icon = "✅" if ok else "❌"
        wrong_line = "" if ok else f'<br><span style="color:#11998e">✓ Correct: {q["answer"]}</span>'
        st.markdown(f"""<div style="border-left:4px solid {color};padding:0.8rem 1rem;margin:0.4rem 0;background:{bg};border-radius:6px;">
            <b>{icon} Q{i+1}. {q['question']}</b>
            <br><span style="color:{color}">Your answer: {ua}</span>{wrong_line}
        </div>""", unsafe_allow_html=True)

# ── PAGE: MY RESULTS ──────────────────────────────────────────────────────────
def show_my_results():
    st.markdown("## 📊 My Quiz History")
    excel_path = get_excel_path(st.session_state.username)
    if not os.path.exists(excel_path):
        st.info("No quiz attempts yet. Take a Pre Quiz or Post Quiz from the sidebar.")
        return
    try:
        df = pd.read_excel(excel_path, sheet_name="Summary")
        st.dataframe(df, use_container_width=True, hide_index=True)
        if len(df) >= 2:
            st.markdown("### 📈 Score Trend")
            fig, ax = plt.subplots(figsize=(10, 4))
            for qt, color, marker in [("Pre Quiz","#2a5298","o"),("Post Quiz","#11998e","s")]:
                sub = df[df["Quiz Type"] == qt]
                if len(sub):
                    ax.plot(range(1, len(sub)+1), sub["Score"], marker=marker, linestyle="-",
                            label=qt, color=color, linewidth=2, markersize=8)
            ax.axhline(y=PASS_MARK, color="red", linestyle="--", alpha=0.7, label=f"Pass Mark ({PASS_MARK})")
            ax.set_xlabel("Attempt Number"); ax.set_ylabel("Score"); ax.set_ylim(0, 26)
            ax.set_title("Your Quiz Score History"); ax.legend(); ax.grid(True, alpha=0.3)
            st.pyplot(fig)
        with open(excel_path, "rb") as f:
            st.download_button("📥 Download My Results Workbook", f.read(),
                f"{st.session_state.username}_quiz_results.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        st.error(f"Error: {e}")

# ── PAGE: INSTRUCTOR DASHBOARD ────────────────────────────────────────────────
def show_instructor_dashboard():
    st.markdown("## 🎓 Instructor Dashboard")
    try:
        df = pd.read_csv(RESULTS_FILE)
    except:
        st.warning("No results yet."); return
    if df.empty:
        st.warning("No results yet."); return

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Students", df["username"].nunique())
    c2.metric("Total Attempts", len(df))
    c3.metric("Average Score", f"{df['score'].mean():.1f}/25")
    pass_rate = (df["result"] == "Pass").mean() * 100 if "result" in df.columns else 0
    c4.metric("Pass Rate", f"{pass_rate:.1f}%")
    st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("### Score Distribution")
        fig, ax = plt.subplots()
        ax.hist(df["score"], bins=range(0,27), color="#2a5298", alpha=0.85, edgecolor="white")
        ax.axvline(x=PASS_MARK, color="red", linestyle="--", label=f"Pass Mark ({PASS_MARK})")
        ax.set_xlabel("Score"); ax.set_ylabel("Count"); ax.legend()
        st.pyplot(fig)
    with col2:
        st.markdown("### Pre vs Post Quiz Averages")
        if "quiz_type" in df.columns:
            avg = df.groupby("quiz_type")["score"].mean().reset_index()
            fig2, ax2 = plt.subplots()
            bars = ax2.bar(avg["quiz_type"], avg["score"], color=["#2a5298","#11998e"], width=0.5)
            ax2.bar_label(bars, fmt="%.1f", padding=5)
            ax2.set_ylim(0, 26)
            ax2.axhline(y=PASS_MARK, color="red", linestyle="--", alpha=0.7, label="Pass Mark")
            ax2.legend(); ax2.set_title("Average by Quiz Type")
            st.pyplot(fig2)

    st.markdown("### All Results")
    sort_col = "date" if "date" in df.columns else "score"
    st.dataframe(df.sort_values(sort_col, ascending=False), use_container_width=True, hide_index=True)
    st.download_button("📥 Download All Results (CSV)", df.to_csv(index=False), "all_results.csv", "text/csv")

# ── PAGE: AI TUTOR ────────────────────────────────────────────────────────────
def show_ai_tutor():
    st.markdown("## 🤖 AI Tutor")
    st.markdown("Ask any Machine Learning question and get an expert explanation.")
    question = st.text_area("Your Question:", placeholder="e.g. What is the difference between bagging and boosting?", height=100)
    if st.button("💬 Ask AI Tutor", type="primary"):
        if question.strip():
            try:
                client = get_openai_client()
                with st.spinner("Thinking..."):
                    resp = client.chat.completions.create(
                        model="gpt-4o-mini",
                        messages=[
                            {"role":"system","content":"You are an expert Machine Learning tutor. Explain clearly with examples. Use markdown headings and bullet points."},
                            {"role":"user","content":question}
                        ]
                    )
                st.markdown("### Answer")
                st.markdown(resp.choices[0].message.content)
            except Exception as e:
                st.error(f"AI Tutor unavailable: {e}")
        else:
            st.warning("Please enter a question.")

# ── ROUTING ───────────────────────────────────────────────────────────────────
if not st.session_state.logged_in:
    show_login()
else:
    with st.sidebar:
        st.markdown(f"### 👤 {st.session_state.name}")
        st.caption(f"Batch: {st.session_state.batch}  |  ID: {st.session_state.username}")
        st.markdown("---")
        menu = ["🏠 Home", "📚 Course Content", "📝 Pre Quiz", "✅ Post Quiz", "📊 My Results"]
        if is_instructor(st.session_state.username):
            menu.append("🎓 Instructor Dashboard")
        menu.append("🤖 AI Tutor")
        page = st.selectbox("Navigation", menu, label_visibility="collapsed")
        st.markdown("---")
        if st.button("🚪 Logout", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()

    if page == "🏠 Home":
        show_home()
    elif page == "📚 Course Content":
        show_course_content()
    elif page == "📝 Pre Quiz":
        show_quiz_page("Pre Quiz")
    elif page == "✅ Post Quiz":
        show_quiz_page("Post Quiz")
    elif page == "📊 My Results":
        show_my_results()
    elif page == "🎓 Instructor Dashboard":
        show_instructor_dashboard()
    elif page == "🤖 AI Tutor":
        show_ai_tutor()
